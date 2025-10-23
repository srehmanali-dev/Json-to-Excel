import os
import sys
import re
import json
import time
import argparse
from pathlib import Path
from typing import Any, Dict, List, Set

import pandas as pd

# -------------------------------
# Utilities
# -------------------------------
INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'

def sanitize_sheet_name(name: str) -> str:
    """Excel sheet name rules: <=31 chars, no : \ / ? * [ ] and not empty."""
    name = re.sub(INVALID_SHEET_CHARS, "_", (name or "").strip())
    name = name or "Sheet"
    return name[:31] if len(name) > 31 else name

def uniquify(name: str, used: Set[str]) -> str:
    base = name
    i = 2
    while name in used:
        suffix = f'_{i}'
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name

def convert_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert ISO-like datetime strings into real datetime columns wherever possible.
    Only replace if >=50% of the column converts. Uses a quick prefilter to skip obvious non-dates.
    """
    for col in df.columns:
        if df[col].dtype == object:
            s = df[col]
            # Prefilter: contains digits and a date-ish delimiter or 'T'
            if not s.dropna().astype(str).str.contains(r'(\d.*[-/T])|([:/-].*\d)', regex=True).any():
                continue
            converted = pd.to_datetime(s, errors="coerce", utc=False, cache=True)
            if converted.notna().sum() >= len(df) / 2:
                df[col] = converted
    return df

def is_array_of_dicts(x: Any) -> bool:
    return isinstance(x, list) and len(x) > 0 and all(isinstance(i, dict) for i in x)

def collect_tables(obj: Any, path: str, out: Dict[str, Any], include_empty: bool) -> None:
    """
    Find arrays-of-objects anywhere in the JSON and store as {path: rows}.
    If include_empty is True, zero-length lists are kept as empty tables (for schema visibility).
    """
    if is_array_of_dicts(obj):
        out[path] = obj
        return

    if isinstance(obj, dict):
        for k, v in obj.items():
            subpath = f"{path}.{k}" if path else k
            if is_array_of_dicts(v):
                out[subpath] = v
            elif isinstance(v, (dict, list)):
                collect_tables(v, subpath, out, include_empty)

    elif isinstance(obj, list):
        if len(obj) == 0 and include_empty:
            out[path] = []  # zero-row table
            return
        for idx, item in enumerate(obj):
            subpath = f"{path}[{idx}]"
            if is_array_of_dicts(item):
                out[subpath] = item
            elif isinstance(item, (dict, list)):
                collect_tables(item, subpath, out, include_empty)

def rows_to_dataframe(rows: Any) -> pd.DataFrame:
    """
    Robust conversion:
    - list[dict] -> DataFrame(rows)
    - []         -> empty DataFrame
    - list[scalar] -> single 'value' column
    - other/mixed -> json_normalize best-effort
    """
    if isinstance(rows, list):
        if len(rows) == 0:
            return pd.DataFrame()
        if all(isinstance(i, dict) for i in rows):
            return pd.DataFrame(rows)
        if all(not isinstance(i, (dict, list)) for i in rows):
            return pd.DataFrame({"value": rows})
        try:
            return pd.json_normalize(rows, sep='.')
        except Exception:
            return pd.DataFrame({"value": [rows]})
    elif isinstance(rows, dict):
        return pd.DataFrame([rows])
    else:
        return pd.DataFrame({"value": [rows]})

def autoformat_sheet(ws):
    # Autofit (approximate): compute max length per column
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column
    max_row = ws.max_row
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, min(max_row, 2000) + 1):  # cap scan for perf
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is None:
                continue
            txt = str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

# -------------------------------
# Main
# -------------------------------
def main():
    parser = argparse.ArgumentParser(description="Extract tables from JSON to Excel or CSV.")
    parser.add_argument("json_file", help="Path to JSON file")
    parser.add_argument("--out", default=None, help="Output directory (default: ./bin/Debug next to script)")
    parser.add_argument("--base-name", default="employee_data", help="Base name for output file(s)")
    parser.add_argument("--sheet-naming", choices=["last", "full"], default="last",
                        help="Use last token or full dot/bracket path for sheet names (trimmed to Excel limits)")
    parser.add_argument("--include-empty", action="store_true",
                        help="Create sheets for empty arrays (zero rows) so schema presence is visible")
    parser.add_argument("--autofit", action="store_true", help="Autofit Excel columns (approximate)")
    parser.add_argument("--freeze", action="store_true", help="Freeze header row in Excel")
    parser.add_argument("--filters", action="store_true", help="Enable Excel autofilter on header row")
    parser.add_argument("--csv", action="store_true",
                        help="Write CSV instead of Excel; if multiple tables are found, writes one CSV per table")
    args = parser.parse_args()

    json_path = Path(args.json_file)
    if not json_path.exists():
        sys.exit(f"❌ JSON file not found: {json_path}")

    try:
        with json_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        sys.exit(f"❌ Invalid JSON format: {e}")

    # Output directory
    script_dir = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()
    out_dir = Path(args.out) if args.out else script_dir / "bin" / "Debug"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Extract tables
    tables: Dict[str, Any] = {}
    collect_tables(data, path="root", out=tables, include_empty=args.include_empty)

    # Fallbacks if nothing found
    if not tables:
        if isinstance(data, list):
            tables["root"] = data
        elif isinstance(data, dict):
            tables["root"] = [data]
        else:
            sys.exit("❌ No arrays found to tabulate in JSON")

    timestamp = time.strftime("%Y%m%d_%H%M%S")

    # CSV branch
    if args.csv:
        used_names: Set[str] = set()

        def short_name(raw_path: str) -> str:
            if args.sheet_naming == "last":
                if ".Result" in raw_path:
                    base = "Result"
                else:
                    m = re.findall(r'([^\.\[\]]+)(?:\[\d+\])?$', raw_path)
                    base = m[0] if m else raw_path
            else:
                base = raw_path
            return uniquify(sanitize_sheet_name(base), used_names)

        if len(tables) == 1:
            (_, rows), = tables.items()
            df = rows_to_dataframe(rows)
            df = convert_datetime_columns(df)
            csv_path = out_dir / f"{args.base_name}_{timestamp}.csv"
            df.to_csv(csv_path, index=False)
            print(f"✅ CSV created at:\n{csv_path}")
        else:
            created: List[Path] = []
            for raw_path, rows in tables.items():
                df = rows_to_dataframe(rows)
                df = convert_datetime_columns(df)
                name = short_name(raw_path)
                csv_path = out_dir / f"{args.base_name}_{name}_{timestamp}.csv"
                df.to_csv(csv_path, index=False)
                created.append(csv_path)
            print(f"✅ {len(created)} CSV file(s) created:")
            for p in created:
                print(p)
        sys.exit(0)

    # Excel branch
    xlsx_path = out_dir / f"{args.base_name}_{timestamp}.xlsx"
    used_names: Set[str] = set()
    summary_rows: List[Dict[str, Any]] = []

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for raw_path, rows in tables.items():
            # Sheet naming
            if args.sheet_naming == "last":
                if ".Result" in raw_path:
                    base_name = "Result"
                else:
                    m = re.findall(r'([^\.\[\]]+)(?:\[\d+\])?$', raw_path)
                    base_name = m[0] if m else raw_path
            else:
                base_name = raw_path.replace("root.", "root.")

            sheet_name = uniquify(sanitize_sheet_name(base_name), used_names)

            # Build DataFrame
            df = rows_to_dataframe(rows)
            df = convert_datetime_columns(df)

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Optional niceties
            ws = writer.sheets[sheet_name]
            if args.freeze and df.shape[0] > 0:
                ws.freeze_panes = "A2"
            if args.filters and df.shape[0] > 0 and df.shape[1] > 0:
                ws.auto_filter.ref = ws.dimensions
            if args.autofit:
                autoformat_sheet(ws)

            summary_rows.append({
                "sheet": sheet_name,
                "rows": int(df.shape[0]),
                "cols": int(df.shape[1]),
                "source_path": raw_path
            })

        # Summary tab last
        summary_df = pd.DataFrame(summary_rows).sort_values(["sheet"])
        summary_df.to_excel(writer, sheet_name="__summary__", index=False)
        ws_sum = writer.sheets["__summary__"]
        if args.freeze and not summary_df.empty:
            ws_sum.freeze_panes = "A2"
        if args.filters and not summary_df.empty:
            ws_sum.auto_filter.ref = ws_sum.dimensions
        if args.autofit:
            autoformat_sheet(ws_sum)

    print(f"✅ Excel file with {len(tables)} sheet(s) created at:\n{xlsx_path}")

if __name__ == "__main__":
    main()
